import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from aliyunapi.monitor import CloudMonitor
from aliyunapi.ecs import check_disk_usage_exceed
from aliyunapi.rds import check_rds_cpu_exceed
from aliyunapi.redis import check_redis_mem_exceed, check_redis_backup
from aliyunapi.ram import check_ram_inactive
from aliyunapi.vpc import check_open_ports, check_unbound_eip

from alibabacloud_tea_openapi.models import Config
from alibabacloud_r_kvstore20150101.client import Client as RedisClient

from utils import get_time_range


class ReportDaily:
    def __init__(self, config_path="config.json"):
        with open(config_path, 'r') as f:
            self.config = json.load(f)
        self.file_path = self.config["FilePath"]

        self.monitor = CloudMonitor(
            self.config["AccessKeyID"],
            self.config["AccessKeySecret"],
            self.config["Region"][0]
        )
        self.start_time = self.config.get("StartTime")
        self.end_time = self.config.get("EndTime")
        if not self.start_time or not self.end_time:
            self.start_time, self.end_time = get_time_range()

    def run(self):
        ecs_data = check_disk_usage_exceed(self.monitor, self.start_time, self.end_time)
        rds_data = check_rds_cpu_exceed(self.monitor, self.start_time, self.end_time)

        # Redis 客户端初始化（新版 SDK）
        redis_config = Config(
            access_key_id=self.config["AccessKeyID"],
            access_key_secret=self.config["AccessKeySecret"],
            region_id=self.config["Region"][0]
        )
        redis_client = RedisClient(redis_config)

        redis_mem_data = check_redis_mem_exceed(self.monitor, redis_client, self.start_time, self.end_time)
        redis_backup_data = check_redis_backup(redis_client)

        ram_data = check_ram_inactive()

        open_port_data = check_open_ports(
            self.config["AccessKeyID"],
            self.config["AccessKeySecret"],
            self.config["Region"][0]
        )
        eip_data = check_unbound_eip(
            self.config["AccessKeyID"],
            self.config["AccessKeySecret"],
            self.config["Region"][0]
        )

        summary = [
            ["资源类型", "检查项目", "阈值/说明", "异常数量", "备注"],
            ["ECS", "磁盘使用率", "> 90%", sum(1 for r in ecs_data if str(r[-1]).startswith("异常")), "磁盘超限"],
            ["RDS", "CPU使用率", "> 80%", sum(1 for r in rds_data if str(r[-1]).startswith("异常")), "CPU超限"],
            ["Redis", "内存使用率", "> 85%", sum(1 for r in redis_mem_data if str(r[-1]).startswith("异常")), "内存使用"],
            ["Redis", "备份策略", "必须配置", sum(1 for r in redis_backup_data if str(r[-1]).startswith("异常")), "备份缺失"],
            ["RAM", "账号使用", "未登录/未使用", sum(1 for r in ram_data if str(r[-1]).startswith("异常")), "账号异常"],
            ["安全组", "端口暴露", "含 22/3306 等", sum(1 for r in open_port_data if str(r[-1]).startswith("异常")), "开放高危端口"],
            ["EIP", "未绑定检查", "InstanceId为空", sum(1 for r in eip_data if str(r[-1]).startswith("异常")), "公网IP未用"]
        ]

        return summary, ecs_data, rds_data, redis_mem_data, redis_backup_data, ram_data, open_port_data, eip_data

    def save(self):
        summary, ecs_data, rds_data, redis_mem_data, redis_backup_data, ram_data, open_port_data, eip_data = self.run()
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "巡检汇总"

        def write_sheet(ws, rows):
            bold_font = Font(bold=True)
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            alignment = Alignment(horizontal="center", vertical="center")

            for i, row in enumerate(rows):
                if ws.title != "巡检汇总" and i > 0:
                    row = [i] + row
                ws.append(row)

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = alignment
                    cell.border = border
                if row[0].row == 1:
                    for cell in row:
                        cell.font = bold_font

            ws.column_dimensions["A"].width = 5
            for col in range(2, len(rows[0]) + 2):
                ws.column_dimensions[chr(64 + col)].width = 20

        # 写入各工作表
        write_sheet(ws_summary, summary)
        write_sheet(wb.create_sheet("ECS 磁盘使用率"), [["序号", "资源类型", "检查项", "实例ID", "使用率", "状态"]] + ecs_data)
        write_sheet(wb.create_sheet("RDS CPU使用率"), [["序号", "资源类型", "检查项", "实例ID", "使用率", "状态"]] + rds_data)
        write_sheet(wb.create_sheet("Redis 内存使用率"), [["序号", "资源类型", "检查项", "实例ID", "使用率", "状态"]] + redis_mem_data)
        write_sheet(wb.create_sheet("Redis 备份检查"), [["序号", "资源类型", "检查项", "实例ID", "是否有备份", "状态"]] + redis_backup_data)
        write_sheet(wb.create_sheet("RAM 异常用户行为"), [["序号", "用户名", "是否已登录", "AK是否使用", "状态"]] + ram_data)
        write_sheet(wb.create_sheet("安全组端口暴露"), [["序号", "资源类型", "检查项", "安全组ID", "协议", "端口", "IP段", "状态"]] + open_port_data)
        write_sheet(wb.create_sheet("EIP 绑定检查"), [["序号", "资源类型", "检查项", "EIP地址", "带宽", "绑定状态", "状态"]] + eip_data)

        wb.save(self.file_path)
        print(f"✅ 巡检报告已保存至：{self.file_path}")
